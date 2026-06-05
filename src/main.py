from pathlib import Path
from datetime import datetime, timedelta
from hashlib import sha1
import json
from unicodedata import normalize
from urllib.parse import quote_plus

import shutil
import pandas as pd
from openpyxl import load_workbook
from playwright.sync_api import TimeoutError as PlaywrightTimeoutError
from playwright.sync_api import sync_playwright

from src import storage as storage_backend


EXPORT_PATH = Path("exports/leads.xlsx")
ACTIVE_EXPORT_PATH = Path("exports/leads_ativos.xlsx")
HISTORY_EXPORT_PATH = Path("exports/leads_historico_completo.xlsx")
FEEDBACKS_PATH = Path("exports/feedbacks.json")
CURRENT_LEADS_PATH = Path("exports/leads_atuais.json")
MASTER_LEADS_PATH = Path("exports/leads_master.json")
RECENT_ACTIONS_PATH = Path("exports/acoes_recentes.json")
SCRAPES_HISTORY_PATH = Path("exports/raspagens.json")
BACKUP_DIR = Path("exports/backups")
LAST_EXPORT_PATH = EXPORT_PATH
LAST_ACTIVE_EXPORT_PATH = ACTIVE_EXPORT_PATH
LAST_HISTORY_EXPORT_PATH = HISTORY_EXPORT_PATH
COUNTRY_CODE = "55"
MAX_COMPANIES = 100
SCROLL_ATTEMPTS = 8
EXPORT_COLUMNS = [
    "Nome",
    "Telefone",
    "WhatsApp",
    "Endereco",
    "Site",
    "Nota",
    "Quantidade de avaliacoes",
    "Cidade",
    "Tem Site?",
    "Oportunidade",
    "Link do Google Maps",
]
FEEDBACK_COLUMNS = [
    "Status abordagem",
    "WhatsApp valido?",
    "Mensagem enviada?",
    "Observacao",
    "Data/hora do feedback",
    "Data ultimo feedback",
    "Data primeira abordagem",
    "Ultima acao",
    "Origem raspagem",
]
FULL_EXPORT_COLUMNS = EXPORT_COLUMNS + FEEDBACK_COLUMNS
PIPELINE_STATUSES = ["NOVO", "SUCESSO", "BURN"]
BRAZILIAN_STATES = {
    "AC": "acre",
    "AL": "alagoas",
    "AP": "amapa",
    "AM": "amazonas",
    "BA": "bahia",
    "CE": "ceara",
    "DF": "distrito federal",
    "ES": "espirito santo",
    "GO": "goias",
    "MA": "maranhao",
    "MT": "mato grosso",
    "MS": "mato grosso do sul",
    "MG": "minas gerais",
    "PA": "para",
    "PB": "paraiba",
    "PR": "parana",
    "PE": "pernambuco",
    "PI": "piaui",
    "RJ": "rio de janeiro",
    "RN": "rio grande do norte",
    "RS": "rio grande do sul",
    "RO": "rondonia",
    "RR": "roraima",
    "SC": "santa catarina",
    "SP": "sao paulo",
    "SE": "sergipe",
    "TO": "tocantins",
}


def normalize_text(value: str) -> str:
    """Remove acentos e padroniza texto para comparacoes mais confiaveis."""
    without_accents = normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return " ".join(without_accents.lower().replace(",", " ").split())


def clean_digits(value: str) -> str:
    """Mantem apenas numeros em telefones e textos parecidos."""
    return "".join(char for char in value if char.isdigit())


def to_int(value: str) -> int:
    """Converte textos numericos vazios ou formatados para inteiro."""
    digits = clean_digits(value)
    return int(digits) if digits else 0


def build_whatsapp_link(phone: str) -> str:
    """Gera o link do WhatsApp usando DDI do Brasil."""
    digits = clean_digits(phone)

    if not digits:
        return ""

    if digits.startswith(COUNTRY_CODE):
        return f"https://wa.me/{digits}"

    return f"https://wa.me/{COUNTRY_CODE}{digits}"


def calculate_opportunity(site: str, reviews: str) -> str:
    """Classifica a oportunidade pelo site e volume de avaliacoes."""
    reviews_count = to_int(reviews)

    if not site and reviews_count <= 20:
        return "ALTA"

    if not site and reviews_count <= 50:
        return "MEDIA"

    return "BAIXA"


def lead_key(lead: dict) -> str:
    """Cria a chave de duplicidade: telefone, Maps ou nome+cidade."""
    phone = clean_digits(lead.get("Telefone", ""))
    maps_link = lead.get("Link do Google Maps", "")
    name = lead.get("Nome", "")
    city = lead.get("Cidade", "")

    if name or city:
        return phone or maps_link or normalize_text(f"{name} {city}")

    return phone or maps_link


def lead_identity_keys(lead: dict) -> list[str]:
    """Lista todas as chaves possiveis para bloquear duplicados."""
    keys = []
    phone = clean_digits(lead.get("Telefone", ""))
    maps_link = lead.get("Link do Google Maps", "").strip()
    name = lead.get("Nome", "")
    city = lead.get("Cidade", "")
    name_city = normalize_text(f"{name} {city}") if name or city else ""

    if phone:
        keys.append(f"phone:{phone}")

    if maps_link:
        keys.append(f"maps:{maps_link}")

    if name_city:
        keys.append(f"namecity:{name_city}")

    return keys


def lead_identifier(lead: dict) -> str:
    """Identifica um lead de forma estavel para feedbacks e base mestra."""
    raw_identifier = lead_key(lead)
    return sha1(raw_identifier.encode("utf-8")).hexdigest()


def default_feedback() -> dict:
    """Estado padrao de abordagem para um lead ainda nao tratado."""
    return {
        "Status abordagem": "NOVO",
        "WhatsApp valido?": "",
        "Mensagem enviada?": "",
        "Observacao": "",
        "Data/hora do feedback": "",
        "Data primeira abordagem": "",
        "Ultima acao": "",
    }


def load_json(path: Path, default):
    """Carrega JSON local sem quebrar se o arquivo ainda nao existir."""
    if not path.exists():
        return default

    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default


def save_json(path: Path, data) -> None:
    """Salva JSON local em exports."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def load_feedbacks() -> dict:
    """Carrega feedbacks salvos em exports/feedbacks.json."""
    return storage_backend.load_feedbacks()


def save_feedbacks(feedbacks: dict) -> None:
    """Salva feedbacks em exports/feedbacks.json."""
    storage_backend.save_feedbacks(feedbacks)


def save_current_leads(leads: list[dict]) -> None:
    """Salva os leads da ultima busca para atualizar planilhas sem banco."""
    save_json(CURRENT_LEADS_PATH, leads)


def load_current_leads() -> list[dict]:
    """Carrega os leads da ultima busca."""
    return load_json(CURRENT_LEADS_PATH, [])


def load_master_leads() -> list[dict]:
    """Carrega a base mestra acumulativa."""
    if storage_backend.get_storage() == "supabase":
        return storage_backend.load_master_leads()

    master = load_json(MASTER_LEADS_PATH, [])

    if master:
        return master

    legacy_leads = load_current_leads()

    if legacy_leads:
        migrated = [normalize_master_lead(lead) for lead in legacy_leads]
        save_master_leads(migrated)
        return migrated

    return []


def save_master_leads(leads: list[dict]) -> None:
    """Salva a base mestra acumulativa."""
    storage_backend.save_master_leads(leads)


def normalize_master_lead(lead: dict) -> dict:
    """Garante campos de controle na base mestra."""
    normalized = dict(lead)
    lead_id = normalized.get("Lead ID") or lead_identifier(normalized)
    normalized["Lead ID"] = lead_id
    normalized["status_abordagem"] = normalized.get("status_abordagem", "NOVO")
    normalized["data_primeira_abordagem"] = normalized.get("data_primeira_abordagem", "")
    normalized["ultima_acao"] = normalized.get("ultima_acao", "")
    normalized["data_ultimo_feedback"] = normalized.get(
        "data_ultimo_feedback",
        normalized.get("Data/hora do feedback", ""),
    )
    normalized["origem_raspagem"] = normalized.get("origem_raspagem", "")
    normalized["Adicionado em"] = normalized.get(
        "Adicionado em",
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    )
    return normalized


def build_lead_index(leads: list[dict]) -> dict:
    """Cria indice por telefone, Maps e nome+cidade."""
    index = {}

    for position, lead in enumerate(leads):
        for key in lead_identity_keys(lead):
            index.setdefault(key, position)

    return index


def deduplicate_leads(leads: list[dict]) -> list[dict]:
    """Remove duplicados internos mantendo o primeiro registro."""
    unique = []
    index = {}

    for lead in leads:
        normalized = normalize_master_lead(lead)
        keys = lead_identity_keys(normalized)

        if not keys or any(key in index for key in keys):
            continue

        index.update({key: len(unique) for key in keys})
        unique.append(normalized)

    return unique


def backup_master_leads() -> Path | None:
    """Cria backup da base mestra antes de importacoes."""
    if not MASTER_LEADS_PATH.exists():
        return None

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    backup_path = BACKUP_DIR / f"leads_master_backup_{timestamp}.json"
    counter = 2

    while backup_path.exists():
        backup_path = BACKUP_DIR / f"leads_master_backup_{timestamp}_{counter}.json"
        counter += 1

    shutil.copyfile(MASTER_LEADS_PATH, backup_path)
    return backup_path


def backup_before_reset() -> Path | None:
    """Cria backup antes de limpar a base mestra."""
    if not MASTER_LEADS_PATH.exists():
        return None

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    backup_path = BACKUP_DIR / f"reset_backup_{timestamp}.json"
    counter = 2

    while backup_path.exists():
        backup_path = BACKUP_DIR / f"reset_backup_{timestamp}_{counter}.json"
        counter += 1

    shutil.copyfile(MASTER_LEADS_PATH, backup_path)
    return backup_path


def reset_lead_base() -> dict:
    """Limpa base, feedbacks e planilhas derivadas sem apagar backups ou historicos brutos."""
    backup_path = backup_before_reset()
    save_json(MASTER_LEADS_PATH, [])
    save_json(FEEDBACKS_PATH, {})

    for path in [ACTIVE_EXPORT_PATH, HISTORY_EXPORT_PATH]:
        try:
            if path.exists():
                path.unlink()
        except PermissionError:
            pass

    return {
        "backup": str(backup_path) if backup_path else "",
        "message": "Base resetada com sucesso.",
    }


def load_scrapes_history() -> list[dict]:
    """Carrega historico simples de raspagens."""
    return storage_backend.load_raspagens()


def save_scrapes_history(history: list[dict]) -> None:
    """Salva historico de raspagens."""
    if storage_backend.get_storage() == "local":
        save_json(SCRAPES_HISTORY_PATH, history)


def register_scrape_history(summary: dict) -> None:
    """Registra a ultima raspagem para auditoria no painel."""
    storage_backend.save_raspagem(summary)


def upsert_leads_from_scrape(
    new_leads: list[dict],
    scrape_info: dict | None = None,
    diagnostico: bool = True,
) -> tuple[list[dict], dict]:
    """Porta unica de entrada: adiciona apenas contatos realmente novos."""
    master = [normalize_master_lead(lead) for lead in load_master_leads()]
    index = build_lead_index(master)
    added = 0
    duplicates = 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    scrape_info = scrape_info or {}
    origin = scrape_info.get("origem_raspagem") or f"{scrape_info.get('nicho', '')} | {scrape_info.get('cidade', '')}".strip(" |")

    for lead in new_leads:
        normalized = normalize_master_lead(lead)
        keys = lead_identity_keys(normalized)

        if not keys:
            duplicates += 1
            if diagnostico:
                print(f"Lead duplicado ignorado: {normalized.get('Nome', 'Sem nome')} | sem identificador")
            continue

        existing_position = next((index[key] for key in keys if key in index), None)

        if existing_position is not None:
            duplicates += 1
            if diagnostico:
                print(f"Lead duplicado ignorado: {normalized.get('Nome', 'Sem nome')}")
            continue

        normalized["status_abordagem"] = "NOVO"
        normalized["Adicionado em"] = now
        normalized["origem_raspagem"] = origin
        master.append(normalized)
        new_position = len(master) - 1

        for key in keys:
            index[key] = new_position

        added += 1

        if diagnostico:
            print(f"Novo lead adicionado: {normalized.get('Nome', 'Sem nome')}")

    save_master_leads(master)

    summary = {
        "data_hora": now,
        "nicho": scrape_info.get("nicho", ""),
        "cidade": scrape_info.get("cidade", ""),
        "limite": scrape_info.get("limite", ""),
        "avaliacoes_maximas": scrape_info.get("avaliacoes_maximas", ""),
        "leads_encontrados": len(new_leads),
        "novos_adicionados": added,
        "duplicados_ignorados": duplicates,
        "total_geral_base": len(master),
    }
    register_scrape_history(summary)
    return master, summary


def merge_into_master(new_leads: list[dict], diagnostico: bool = True) -> tuple[list[dict], dict]:
    """Compatibilidade: usa a porta unica de entrada."""
    return upsert_leads_from_scrape(new_leads, {}, diagnostico)


def normalize_header(value) -> str:
    """Normaliza nomes de colunas importadas do Excel."""
    return normalize_text(str(value or ""))


def get_row_color_status(row) -> str:
    """Detecta amarelo como SUCESSO, vermelho como BURN e sem cor como NOVO."""
    has_yellow = False

    for cell in row:
        fill = cell.fill

        if not fill or fill.fill_type is None:
            continue

        colors = [fill.fgColor, fill.start_color]

        for color in colors:
            rgb = color.rgb

            if not isinstance(rgb, str) or len(rgb) < 6:
                continue

            rgb = rgb[-6:].upper()
            red = int(rgb[0:2], 16)
            green = int(rgb[2:4], 16)
            blue = int(rgb[4:6], 16)

            if red >= 180 and green < 120 and blue < 120:
                return "BURN"

            if red >= 180 and green >= 150 and blue < 140:
                has_yellow = True

    return "SUCESSO" if has_yellow else "NOVO"


def lead_from_imported_row(headers: dict, values: dict) -> dict:
    """Monta um lead a partir de uma linha importada."""
    def cell_to_text(value) -> str:
        if value is None:
            return ""

        if isinstance(value, float) and value.is_integer():
            return str(int(value))

        return str(value).strip()

    def pick(*names):
        for name in names:
            key = normalize_header(name)
            value = values.get(key, "")
            text = cell_to_text(value)

            if text:
                return text

        return ""

    phone = pick("Telefone", "Phone", "Celular")
    site = pick("Site", "Website")
    reviews = pick("Quantidade de avaliacoes", "Avaliacoes")

    lead = {
        "Nome": pick("Nome", "Empresa", "Lead"),
        "Telefone": phone,
        "WhatsApp": pick("WhatsApp") or build_whatsapp_link(phone),
        "Endereco": pick("Endereco", "Address"),
        "Site": site,
        "Nota": pick("Nota", "Rating"),
        "Quantidade de avaliacoes": reviews,
        "Cidade": pick("Cidade", "City"),
        "Tem Site?": pick("Tem Site?", "Tem Site") or ("SIM" if site else "NAO"),
        "Oportunidade": pick("Oportunidade") or calculate_opportunity(site, reviews),
        "Link do Google Maps": pick("Link do Google Maps", "Link Google Maps", "Google Maps", "Maps"),
    }
    return lead


def import_feedback_spreadsheet(file_path: str | Path) -> dict:
    """Importa feedback por cor de uma planilha Excel."""
    backup_path = backup_master_leads()
    workbook = load_workbook(file_path)
    sheet = workbook.active
    rows = list(sheet.iter_rows())

    summary = {
        "total_linhas": 0,
        "sucessos_importados": 0,
        "burns_importados": 0,
        "novos_importados": 0,
        "duplicados_ignorados": 0,
        "novos_adicionados": 0,
        "existentes_atualizados": 0,
        "total_final_base": 0,
        "restantes_para_abordar": 0,
        "backup": str(backup_path) if backup_path else "",
    }

    if not rows:
        return summary

    header_row = rows[0]
    header_map = {
        index: normalize_header(cell.value)
        for index, cell in enumerate(header_row)
        if normalize_header(cell.value)
    }
    master = [normalize_master_lead(lead) for lead in load_master_leads()]
    index_by_id = {lead["Lead ID"]: index for index, lead in enumerate(master)}
    feedbacks = load_feedbacks()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for row in rows[1:]:
        if all(cell.value in [None, ""] for cell in row):
            continue

        summary["total_linhas"] += 1
        status = get_row_color_status(row)

        values = {
            header: row[index].value
            for index, header in header_map.items()
            if index < len(row)
        }
        lead_data = lead_from_imported_row(header_map, values)

        if not lead_key(lead_data):
            summary["duplicados_ignorados"] += 1
            continue

        lead = normalize_master_lead(lead_data)
        lead_id = lead["Lead ID"]

        if lead_id in index_by_id:
            master[index_by_id[lead_id]]["status_abordagem"] = status
            summary["existentes_atualizados"] += 1
        else:
            lead["status_abordagem"] = status
            lead["Adicionado em"] = now
            index_by_id[lead_id] = len(master)
            master.append(lead)
            summary["novos_adicionados"] += 1

        if status == "NOVO":
            summary["novos_importados"] += 1
            feedbacks.setdefault(lead_id, default_feedback())
        elif status == "SUCESSO":
            summary["sucessos_importados"] += 1
            feedbacks[lead_id] = {
                "Status abordagem": "SUCESSO",
                "WhatsApp valido?": "SIM",
                "Mensagem enviada?": "SIM",
                "Observacao": "Importado da planilha",
                "Data/hora do feedback": now,
                "Data primeira abordagem": feedbacks.get(lead_id, {}).get("Data primeira abordagem", ""),
                "Ultima acao": "Importado da planilha",
            }
        elif status == "BURN":
            summary["burns_importados"] += 1
            feedbacks[lead_id] = {
                "Status abordagem": "BURN",
                "WhatsApp valido?": "NAO",
                "Mensagem enviada?": "NAO",
                "Observacao": "Importado da planilha como BURN",
                "Data/hora do feedback": now,
                "Data primeira abordagem": feedbacks.get(lead_id, {}).get("Data primeira abordagem", ""),
                "Ultima acao": "Importado da planilha",
            }

    save_master_leads(master)
    save_feedbacks(feedbacks)

    if storage_backend.get_storage() == "supabase":
        for feedback_lead_id, feedback in feedbacks.items():
            if feedback.get("Status abordagem") in ["SUCESSO", "BURN"]:
                storage_backend.add_feedback(feedback_lead_id, feedback)

    export_feedback_excels(master)
    stats = prospecting_stats(enrich_leads_with_feedback(master))
    summary["total_final_base"] = stats["total_leads"]
    summary["restantes_para_abordar"] = stats["restantes_para_abordar"]
    return summary


def enrich_leads_with_feedback(leads: list[dict]) -> list[dict]:
    """Adiciona status e observacoes aos leads."""
    feedbacks = load_feedbacks()
    enriched = []

    for lead in leads:
        lead_with_feedback = dict(lead)
        lead_id = lead_with_feedback.get("Lead ID") or lead_identifier(lead_with_feedback)
        lead_with_feedback["Lead ID"] = lead_id
        status = lead_with_feedback.get("status_abordagem", "NOVO")
        feedback = default_feedback()
        feedback["Status abordagem"] = status
        feedback.update(feedbacks.get(lead_id, {}))
        feedback["Data primeira abordagem"] = (
            lead_with_feedback.get("data_primeira_abordagem")
            or feedback.get("Data primeira abordagem", "")
        )
        feedback["Ultima acao"] = (
            lead_with_feedback.get("ultima_acao")
            or feedback.get("Ultima acao", "")
        )
        lead_with_feedback.update(feedback)
        lead_with_feedback["status_abordagem"] = lead_with_feedback["Status abordagem"]
        lead_with_feedback["Data ultimo feedback"] = (
            lead_with_feedback.get("data_ultimo_feedback")
            or lead_with_feedback.get("Data/hora do feedback", "")
        )
        lead_with_feedback["Origem raspagem"] = lead_with_feedback.get("origem_raspagem", "")
        enriched.append(lead_with_feedback)

    return enriched


def load_recent_actions() -> list[dict]:
    """Carrega historico curto para desfazer a ultima acao da esteira."""
    return storage_backend.load_recent_actions()


def save_recent_actions(actions: list[dict]) -> None:
    """Salva historico curto de acoes da esteira."""
    storage_backend.save_recent_actions(actions)


def snapshot_lead(lead: dict, feedback: dict | None = None) -> dict:
    """Guarda o estado minimo para desfazer uma acao."""
    return {
        "lead": dict(lead),
        "feedback": dict(feedback or {}),
    }


def record_recent_action(action: str, lead_id: str, before: dict, after: dict | None = None) -> None:
    """Registra uma acao reversivel feita na esteira."""
    storage_backend.save_recent_action(
        {
            "data_hora": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "acao": action,
            "lead_id": lead_id,
            "before": before,
            "after": after or {},
        }
    )


def update_lead_status(
    lead_id: str,
    status: str,
    whatsapp_valido: str = "",
    mensagem_enviada: str = "",
    observacao: str = "",
    ultima_acao: str = "Status atualizado",
    set_first_approach: bool = False,
    record_action: bool = True,
) -> tuple[list[dict], dict]:
    """Atualiza status, feedbacks e planilhas em um unico ponto."""
    master = [normalize_master_lead(lead) for lead in load_master_leads()]
    feedbacks = load_feedbacks()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    found = False

    for lead in master:
        if lead["Lead ID"] != lead_id:
            continue

        previous_feedback = feedbacks.get(lead_id, default_feedback())
        before = snapshot_lead(lead, previous_feedback)
        lead["status_abordagem"] = status

        if set_first_approach and not lead.get("data_primeira_abordagem"):
            lead["data_primeira_abordagem"] = now

        lead["ultima_acao"] = ultima_acao
        lead["data_ultimo_feedback"] = now

        feedback = feedbacks.get(lead_id, default_feedback())
        feedback["Status abordagem"] = status
        feedback["WhatsApp valido?"] = whatsapp_valido
        feedback["Mensagem enviada?"] = mensagem_enviada
        feedback["Observacao"] = observacao.strip()
        feedback["Data/hora do feedback"] = now
        feedback["Data ultimo feedback"] = now
        feedback["Data primeira abordagem"] = lead.get("data_primeira_abordagem", "")
        feedback["Ultima acao"] = ultima_acao
        feedbacks[lead_id] = feedback

        if record_action:
            record_recent_action(
                ultima_acao,
                lead_id,
                before,
                snapshot_lead(lead, feedback),
            )

        found = True
        break

    if not found:
        return enrich_leads_with_feedback(master), prospecting_stats(enrich_leads_with_feedback(master))

    save_master_leads(master)
    save_feedbacks(feedbacks)

    if storage_backend.get_storage() == "supabase" and lead_id in feedbacks:
        storage_backend.add_feedback(lead_id, feedbacks[lead_id])

    export_feedback_excels(master)
    enriched = enrich_leads_with_feedback(master)
    return enriched, prospecting_stats(enriched)


def prospecting_stats(leads: list[dict]) -> dict:
    """Calcula contadores simples de abordagem."""
    today = datetime.now().strftime("%Y-%m-%d")
    now = datetime.now()
    week_start = (now - timedelta(days=now.weekday())).date()
    month_start = now.replace(day=1).date()

    def sent_reference_date(lead: dict):
        reference_date = (
            lead.get("Data ultimo feedback")
            or lead.get("data_ultimo_feedback")
            or lead.get("Data/hora do feedback")
            or lead.get("Data primeira abordagem")
            or ""
        )

        try:
            return datetime.strptime(reference_date[:10], "%Y-%m-%d").date()
        except ValueError:
            return None

    total = len(leads)
    new = sum(1 for lead in leads if lead.get("Status abordagem") == "NOVO")
    pending = sum(1 for lead in leads if lead.get("Status abordagem") == "PENDENTE")
    success = sum(1 for lead in leads if lead.get("Status abordagem") == "SUCESSO")
    burn = sum(1 for lead in leads if lead.get("Status abordagem") == "BURN")
    sent_today = sum(
        1
        for lead in leads
        if lead.get("Mensagem enviada?") == "SIM"
        and str(sent_reference_date(lead)) == today
    )
    sent_week = 0
    sent_month = 0

    for lead in leads:
        if lead.get("Mensagem enviada?") != "SIM":
            continue

        parsed_date = sent_reference_date(lead)

        if not parsed_date:
            continue

        if parsed_date >= week_start:
            sent_week += 1

        if parsed_date >= month_start:
            sent_month += 1
    approached_today = sum(
        1
        for lead in leads
        if lead.get("Data primeira abordagem", "").startswith(today)
    )
    success_today = sum(
        1
        for lead in leads
        if lead.get("Status abordagem") == "SUCESSO"
        and lead.get("Data/hora do feedback", "").startswith(today)
    )
    burn_today = sum(
        1
        for lead in leads
        if lead.get("Status abordagem") == "BURN"
        and lead.get("Data/hora do feedback", "").startswith(today)
    )
    pending_today = sum(
        1
        for lead in leads
        if lead.get("Status abordagem") == "PENDENTE"
        and lead.get("Data primeira abordagem", "").startswith(today)
    )
    daily_goal = 50
    remaining_queue = new
    finished_contacts = success + burn
    whatsapp_valid_rate = round((success / finished_contacts) * 100) if finished_contacts else 0
    estimated_days = ""

    if daily_goal:
        estimated_days = (remaining_queue + daily_goal - 1) // daily_goal

    return {
        "total_leads": total,
        "novos": new,
        "pendentes": pending,
        "sucesso": success,
        "burn": burn,
        "mensagens_enviadas_hoje": sent_today,
        "mensagens_enviadas_semana": sent_week,
        "mensagens_enviadas_mes": sent_month,
        "restantes_para_abordar": remaining_queue,
        "abordados_hoje": approached_today,
        "sucesso_hoje": success_today,
        "burn_hoje": burn_today,
        "pendentes_gerados_hoje": pending_today,
        "meta_diaria": daily_goal,
        "faltam_para_meta": max(0, daily_goal - approached_today),
        "estimativa_finalizar": estimated_days,
        "taxa_whatsapp_valido": whatsapp_valid_rate,
    }


def export_active_excel(leads: list[dict]) -> Path:
    """Gera planilha ativa com NOVO e SUCESSO."""
    global LAST_ACTIVE_EXPORT_PATH

    EXPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    enriched = enrich_leads_with_feedback(leads)
    active_leads = [lead for lead in enriched if lead.get("Status abordagem") in ["NOVO", "SUCESSO"]]
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    active_fallback = EXPORT_PATH.parent / f"leads_ativos_{timestamp}.xlsx"

    try:
        pd.DataFrame(active_leads, columns=FULL_EXPORT_COLUMNS).to_excel(
            ACTIVE_EXPORT_PATH,
            index=False,
            engine="openpyxl",
        )
        LAST_ACTIVE_EXPORT_PATH = ACTIVE_EXPORT_PATH
    except PermissionError:
        pd.DataFrame(active_leads, columns=FULL_EXPORT_COLUMNS).to_excel(
            active_fallback,
            index=False,
            engine="openpyxl",
        )
        LAST_ACTIVE_EXPORT_PATH = active_fallback

    return LAST_ACTIVE_EXPORT_PATH


def export_full_history_excel(leads: list[dict]) -> Path:
    """Gera historico completo com todos os status e metadados."""
    global LAST_HISTORY_EXPORT_PATH

    EXPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    enriched = enrich_leads_with_feedback(leads)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    history_fallback = EXPORT_PATH.parent / f"leads_historico_completo_{timestamp}.xlsx"

    try:
        pd.DataFrame(enriched, columns=FULL_EXPORT_COLUMNS).to_excel(
            HISTORY_EXPORT_PATH,
            index=False,
            engine="openpyxl",
        )
        LAST_HISTORY_EXPORT_PATH = HISTORY_EXPORT_PATH
    except PermissionError:
        pd.DataFrame(enriched, columns=FULL_EXPORT_COLUMNS).to_excel(
            history_fallback,
            index=False,
            engine="openpyxl",
        )
        LAST_HISTORY_EXPORT_PATH = history_fallback

    return LAST_HISTORY_EXPORT_PATH


def export_feedback_excels(leads: list[dict]) -> None:
    """Gera planilhas de leads ativos e historico completo com feedback."""
    export_active_excel(leads)
    export_full_history_excel(leads)


def save_feedback(lead_id: str, feedback: dict) -> tuple[list[dict], dict]:
    """Atualiza feedback local e regenera planilhas de prospeccao."""
    requested_status = feedback.get("status", "SUCESSO")
    return update_lead_status(
        lead_id,
        requested_status,
        feedback.get("whatsapp_valido", ""),
        feedback.get("mensagem_enviada", ""),
        feedback.get("observacao", ""),
        feedback.get("ultima_acao", "Feedback salvo"),
        requested_status in ["SUCESSO", "BURN"],
    )


def opportunity_rank(lead: dict) -> int:
    """Ordena oportunidade para a esteira."""
    return {"ALTA": 0, "MEDIA": 1, "BAIXA": 2}.get(lead.get("Oportunidade", "BAIXA"), 3)


def status_rank(lead: dict) -> int:
    """Mantem compatibilidade com status legados."""
    return {"NOVO": 0, "PENDENTE": 1}.get(lead.get("Status abordagem", ""), 9)


def get_next_lead_for_queue(status_filter: str = "NOVO", skipped_ids: list[str] | None = None) -> dict | None:
    """Retorna o proximo lead da fila por status."""
    skipped = set(skipped_ids or [])
    requested_status = (status_filter or "NOVO").upper()
    leads = enrich_leads_with_feedback(load_master_leads())
    queue = [
        lead
        for lead in leads
        if lead.get("Status abordagem") == requested_status
        and lead.get("Lead ID") not in skipped
    ]
    queue.sort(key=lambda lead: (opportunity_rank(lead), lead.get("Adicionado em", "")))
    return queue[0] if queue else None


def get_next_pipeline_lead(skip_id: str = "", status_filter: str = "NOVO") -> dict | None:
    """Compatibilidade para a esteira principal."""
    skipped = [skip_id] if skip_id else []
    return get_next_lead_for_queue(status_filter, skipped)


def mark_pipeline_pending(lead_id: str) -> tuple[dict | None, dict]:
    """Registra abertura do WhatsApp sem mudar o status para PENDENTE."""
    master = [normalize_master_lead(lead) for lead in load_master_leads()]
    feedbacks = load_feedbacks()
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for lead in master:
        if lead["Lead ID"] == lead_id:
            previous_feedback = feedbacks.get(lead_id, default_feedback())
            before = snapshot_lead(lead, previous_feedback)
            if not lead.get("data_primeira_abordagem"):
                lead["data_primeira_abordagem"] = now
            lead["ultima_acao"] = "WhatsApp aberto"
            feedback = feedbacks.get(lead_id, default_feedback())
            feedback["Status abordagem"] = lead.get("status_abordagem", "NOVO")
            feedback["Data primeira abordagem"] = feedback.get("Data primeira abordagem") or now
            feedback["Ultima acao"] = "WhatsApp aberto"
            feedbacks[lead_id] = feedback
            record_recent_action("WhatsApp aberto", lead_id, before, snapshot_lead(lead, feedback))
            break

    save_master_leads(master)
    save_feedbacks(feedbacks)
    export_feedback_excels(master)
    enriched = enrich_leads_with_feedback(master)
    return get_next_pipeline_lead(), prospecting_stats(enriched)


def skip_pipeline_lead(lead_id: str, status_filter: str = "NOVO") -> tuple[dict | None, dict]:
    """Registra contato pulado sem alterar status final."""
    master = [normalize_master_lead(lead) for lead in load_master_leads()]
    feedbacks = load_feedbacks()

    for lead in master:
        if lead["Lead ID"] == lead_id:
            before = snapshot_lead(lead, feedbacks.get(lead_id, default_feedback()))
            lead["ultima_acao"] = "Contato pulado"
            record_recent_action("Contato pulado", lead_id, before, snapshot_lead(lead, feedbacks.get(lead_id, default_feedback())))
            break

    save_master_leads(master)
    export_feedback_excels(master)
    enriched = enrich_leads_with_feedback(master)
    next_lead = get_next_pipeline_lead(lead_id, status_filter)
    return next_lead, prospecting_stats(enriched)


def burn_pipeline_lead(lead_id: str) -> tuple[dict | None, dict]:
    """Marca o lead atual como BURN pela esteira."""
    save_feedback(
        lead_id,
        {
            "status": "BURN",
            "whatsapp_valido": "NAO",
            "mensagem_enviada": "NAO",
            "observacao": "Marcado como BURN pela esteira",
            "ultima_acao": "Marcado como BURN",
        },
    )
    enriched = enrich_leads_with_feedback(load_master_leads())
    return get_next_pipeline_lead(), prospecting_stats(enriched)


def undo_last_pipeline_action() -> tuple[dict | None, dict, str]:
    """Restaura o estado anterior do ultimo lead alterado na esteira."""
    actions = load_recent_actions()

    if not actions:
        enriched = enrich_leads_with_feedback(load_master_leads())
        return get_next_pipeline_lead(), prospecting_stats(enriched), "Nenhuma acao recente para desfazer."

    action = actions.pop()
    lead_id = action.get("lead_id", "")
    previous = action.get("before", {})
    previous_lead = previous.get("lead", {})
    previous_feedback = previous.get("feedback", {})
    master = [normalize_master_lead(lead) for lead in load_master_leads()]
    feedbacks = load_feedbacks()

    for index, lead in enumerate(master):
        if lead["Lead ID"] == lead_id:
            master[index] = normalize_master_lead(previous_lead)
            break

    if previous_feedback:
        feedbacks[lead_id] = previous_feedback
    elif lead_id in feedbacks:
        feedbacks.pop(lead_id)

    save_recent_actions(actions)
    save_master_leads(master)
    save_feedbacks(feedbacks)
    export_feedback_excels(master)
    enriched = enrich_leads_with_feedback(master)
    return get_next_pipeline_lead(), prospecting_stats(enriched), "Ultima acao desfeita."


def get_pipeline_leads(status_filter: str = "NOVO") -> tuple[list[dict], dict]:
    """Retorna a base mestra filtrada para a tela."""
    enriched = enrich_leads_with_feedback(load_master_leads())
    status = (status_filter or "NOVO").upper()

    if status != "TODOS":
        enriched = [lead for lead in enriched if lead.get("Status abordagem") == status]

    all_leads = enrich_leads_with_feedback(load_master_leads())
    return enriched, prospecting_stats(all_leads)


def get_active_export_path() -> Path:
    """Retorna a planilha de leads ativos quando existir."""
    if LAST_ACTIVE_EXPORT_PATH.exists():
        return LAST_ACTIVE_EXPORT_PATH

    master = load_master_leads()

    if master:
        export_feedback_excels(master)

        if LAST_ACTIVE_EXPORT_PATH.exists():
            return LAST_ACTIVE_EXPORT_PATH

    active_files = sorted(
        EXPORT_PATH.parent.glob("leads_ativos*.xlsx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )

    if active_files:
        return active_files[0]

    return get_latest_export_path()


def get_history_export_path() -> Path:
    """Retorna a planilha de historico completo quando existir."""
    if LAST_HISTORY_EXPORT_PATH.exists():
        return LAST_HISTORY_EXPORT_PATH

    master = load_master_leads()

    if master:
        export_feedback_excels(master)

        if LAST_HISTORY_EXPORT_PATH.exists():
            return LAST_HISTORY_EXPORT_PATH

    history_files = sorted(
        EXPORT_PATH.parent.glob("leads_historico_completo*.xlsx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )

    if history_files:
        return history_files[0]

    return get_latest_export_path()


def get_latest_export_path() -> Path:
    """Retorna o ultimo arquivo Excel gerado nesta execucao."""
    if LAST_EXPORT_PATH.exists():
        return LAST_EXPORT_PATH

    timestamped_files = sorted(
        EXPORT_PATH.parent.glob("leads_*.xlsx"),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )

    if timestamped_files:
        return timestamped_files[0]

    return LAST_EXPORT_PATH


def split_city_and_state(city: str) -> tuple[str, str]:
    """Separa entradas como 'Montenegro RS' em cidade e UF."""
    parts = city.strip().split()

    if len(parts) < 2:
        return city.strip(), ""

    possible_state = parts[-1].upper().replace(".", "")

    if possible_state in BRAZILIAN_STATES:
        return " ".join(parts[:-1]).strip(), possible_state

    return city.strip(), ""


def address_matches_city(address: str, searched_city: str) -> bool:
    """Confere se o endereco pertence rigorosamente a cidade pesquisada."""
    city_name, state = split_city_and_state(searched_city)
    normalized_address = normalize_text(address)
    normalized_city = normalize_text(city_name)

    if not normalized_city or normalized_city not in normalized_address:
        return False

    if not state:
        return True

    normalized_state_name = normalize_text(BRAZILIAN_STATES[state])
    normalized_state_code = state.lower()
    address_with_spaces = f" {normalized_address} "

    return f" {normalized_state_code} " in address_with_spaces or normalized_state_name in normalized_address


def safe_text(page, selector: str) -> str:
    """Busca texto na pagina sem interromper o robo quando o campo nao existe."""
    try:
        locator = page.locator(selector).first
        if locator.count() == 0:
            return ""

        return locator.inner_text(timeout=1500).strip()
    except PlaywrightTimeoutError:
        return ""
    except Exception:
        return ""


def safe_attribute(page, selector: str, attribute: str) -> str:
    """Busca atributo HTML na pagina sem interromper a execucao."""
    try:
        locator = page.locator(selector).first
        if locator.count() == 0:
            return ""

        return locator.get_attribute(attribute, timeout=1500) or ""
    except PlaywrightTimeoutError:
        return ""
    except Exception:
        return ""


def extract_rating_and_reviews(page) -> tuple[str, str]:
    """Extrai nota e quantidade de avaliacoes quando o Google Maps mostra esses dados."""
    rating = safe_text(page, "div.F7nice span[aria-hidden='true']")
    reviews_text = safe_text(page, "div.F7nice span[aria-label*='avalia']")

    reviews = clean_digits(reviews_text)
    return rating, reviews


def extract_company_data(page) -> dict:
    """Coleta os dados exibidos no painel lateral da empresa."""
    name = safe_text(page, "h1.DUwDvf")
    phone = safe_text(page, "button[data-item-id^='phone'] div.Io6YTe")
    address = safe_text(page, "button[data-item-id='address'] div.Io6YTe")
    site = safe_attribute(page, "a[data-item-id='authority']", "href")
    maps_link = page.url
    rating, reviews = extract_rating_and_reviews(page)

    return {
        "Nome": name,
        "Cidade": "",
        "Telefone": phone,
        "WhatsApp": build_whatsapp_link(phone),
        "Endereco": address,
        "Site": site,
        "Tem Site?": "SIM" if site else "NAO",
        "Oportunidade": calculate_opportunity(site, reviews),
        "Nota": rating,
        "Quantidade de avaliacoes": reviews,
        "Link do Google Maps": maps_link,
    }


def open_google_maps(page, niche: str, city: str) -> None:
    """Abre o Google Maps ja com a busca preenchida."""
    search = quote_plus(f"{niche} em {city}")
    url = f"https://www.google.com/maps/search/{search}"
    page.goto(url, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(5000)


def collect_result_links(page, limit: int) -> list[str]:
    """Carrega resultados da busca e retorna links unicos de empresas."""
    links = []
    feed = page.locator("div[role='feed']")
    result_limit = max(1, limit)
    max_scrolls = max(SCROLL_ATTEMPTS, result_limit // 5 + 10)

    for _ in range(max_scrolls):
        result_links = page.locator("a.hfpxzc").all()

        for link in result_links:
            href = link.get_attribute("href")
            if href and href not in links:
                links.append(href)

            if len(links) >= result_limit:
                return links

        if feed.count() > 0:
            feed.first.evaluate("element => element.scrollBy(0, element.scrollHeight)")

        page.wait_for_timeout(2000)

    return links


def get_discard_reason(
    company: dict,
    searched_city: str,
    only_with_phone: bool,
    only_without_site: bool,
    max_reviews: int | None = None,
) -> tuple[bool, str, str]:
    """Retorna se a empresa deve ser mantida e o motivo quando descartada."""
    is_in_city = address_matches_city(company["Endereco"], searched_city)
    has_phone = bool(clean_digits(company["Telefone"]))
    has_site = bool(company["Site"])
    reviews_count = to_int(company["Quantidade de avaliacoes"])

    if not is_in_city:
        return False, "cidade_diferente", "fora da cidade pesquisada"

    if only_with_phone and not has_phone:
        return False, "sem_telefone", "sem telefone"

    if only_without_site and has_site:
        return False, "possui_site", "possui site"

    if max_reviews is not None and reviews_count > max_reviews:
        return (
            False,
            "avaliacoes_acima",
            f"{reviews_count} avaliacoes, maximo permitido {max_reviews}",
        )

    return True, "", ""


def should_keep_company(
    company: dict,
    searched_city: str,
    only_with_phone: bool,
    only_without_site: bool,
    max_reviews: int | None = None,
) -> bool:
    """Aplica os filtros escolhidos pelo usuario."""
    keep, _, _ = get_discard_reason(
        company,
        searched_city,
        only_with_phone,
        only_without_site,
        max_reviews,
    )
    return keep


def print_diagnostic(company: dict, accepted: bool, reason: str) -> None:
    """Mostra no terminal o diagnostico da empresa analisada."""
    name = company["Nome"] or "Empresa sem nome"
    has_phone = "SIM" if clean_digits(company["Telefone"]) else "NAO"
    has_site = "SIM" if company["Site"] else "NAO"
    reviews = company["Quantidade de avaliacoes"] or "0"

    details = (
        f"{name} | telefone encontrado {has_phone} | "
        f"site encontrado {has_site} | avaliacoes encontradas {reviews}"
    )

    if accepted:
        print(f"ACEITO: {details}")
    else:
        print(f"DESCARTADO: {details} | Motivo: {reason}")


def export_to_excel(companies: list[dict]) -> dict:
    """Salva Excel datado e tenta atualizar exports/leads.xlsx."""
    global LAST_EXPORT_PATH

    EXPORT_PATH.parent.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    timestamped_path = EXPORT_PATH.parent / f"leads_{timestamp}.xlsx"
    counter = 2

    while timestamped_path.exists():
        timestamped_path = EXPORT_PATH.parent / f"leads_{timestamp}_{counter}.xlsx"
        counter += 1

    df = pd.DataFrame(companies, columns=EXPORT_COLUMNS)
    df.to_excel(timestamped_path, index=False, engine="openpyxl")

    result = {
        "path": timestamped_path,
        "main_path": EXPORT_PATH,
        "main_updated": True,
        "message": "",
    }

    try:
        shutil.copyfile(timestamped_path, EXPORT_PATH)
    except PermissionError:
        result["main_updated"] = False
        result["message"] = (
            "A planilha principal est\u00e1 aberta no Excel. "
            "Geramos uma nova planilha com data e hora."
        )

    LAST_EXPORT_PATH = timestamped_path
    return result


def buscar_leads(
    nicho: str,
    cidade: str,
    limite: int = MAX_COMPANIES,
    apenas_com_telefone: bool = True,
    apenas_sem_site: bool = True,
    avaliacoes_maximas: int | None = None,
    retornar_estatisticas: bool = False,
    diagnostico: bool = True,
) -> list[dict] | tuple[list[dict], dict]:
    """Busca leads no Google Maps, aplica filtros e exporta o Excel."""
    leads = []
    result_limit = max(1, int(limite or MAX_COMPANIES))
    stats = {
        "empresas_analisadas": 0,
        "leads_encontrados": 0,
        "leads_raspagem": 0,
        "novos_adicionados": 0,
        "duplicados_ignorados": 0,
        "total_geral_base": 0,
        "leads_com_telefone": 0,
        "leads_sem_site": 0,
        "descartados_sem_telefone": 0,
        "descartados_possui_site": 0,
        "descartados_avaliacoes": 0,
        "descartados_cidade": 0,
        "total_descartado": 0,
        "arquivo_excel": "",
        "mensagem_exportacao": "",
    }

    with sync_playwright() as playwright:
        browser = playwright.chromium.launch(headless=False)
        try:
            page = browser.new_page(locale="pt-BR")

            open_google_maps(page, nicho, cidade)
            result_links = collect_result_links(page, result_limit)
            stats["empresas_analisadas"] = len(result_links)

            for index, link in enumerate(result_links, start=1):
                print(f"Analisando {index} de {len(result_links)}")
                page.goto(link, wait_until="domcontentloaded", timeout=60000)
                page.wait_for_timeout(2500)

                company = extract_company_data(page)
                company["Cidade"] = cidade
                company["Tem Site?"] = "SIM" if company["Site"] else "NAO"
                company["Oportunidade"] = calculate_opportunity(
                    company["Site"],
                    company["Quantidade de avaliacoes"],
                )

                accepted, reason_key, reason = get_discard_reason(
                    company,
                    cidade,
                    apenas_com_telefone,
                    apenas_sem_site,
                    avaliacoes_maximas,
                )

                if diagnostico:
                    print_diagnostic(company, accepted, reason)

                if accepted:
                    leads.append(company)
                else:
                    stats["total_descartado"] += 1

                    if reason_key == "sem_telefone":
                        stats["descartados_sem_telefone"] += 1
                    elif reason_key == "possui_site":
                        stats["descartados_possui_site"] += 1
                    elif reason_key == "avaliacoes_acima":
                        stats["descartados_avaliacoes"] += 1
                    elif reason_key == "cidade_diferente":
                        stats["descartados_cidade"] += 1
        finally:
            browser.close()

    export_result = export_to_excel(leads)
    master, merge_summary = upsert_leads_from_scrape(
        leads,
        {
            "nicho": nicho,
            "cidade": cidade,
            "limite": result_limit,
            "avaliacoes_maximas": avaliacoes_maximas if avaliacoes_maximas is not None else "",
        },
        diagnostico,
    )
    export_feedback_excels(master)
    enriched_leads = enrich_leads_with_feedback(master)

    stats["leads_raspagem"] = len(leads)
    stats["leads_encontrados"] = len(enriched_leads)
    stats["novos_adicionados"] = merge_summary["novos_adicionados"]
    stats["duplicados_ignorados"] = merge_summary["duplicados_ignorados"]
    stats["total_geral_base"] = merge_summary["total_geral_base"]
    stats["leads_com_telefone"] = sum(1 for lead in enriched_leads if clean_digits(lead["Telefone"]))
    stats["leads_sem_site"] = sum(1 for lead in enriched_leads if not lead["Site"])
    stats["arquivo_excel"] = str(export_result["path"])
    stats["mensagem_exportacao"] = export_result["message"]

    if diagnostico:
        print("\nResumo do diagnostico")
        print(f"Total analisado: {stats['empresas_analisadas']}")
        print(f"Total descartado por sem telefone: {stats['descartados_sem_telefone']}")
        print(f"Total descartado por possuir site: {stats['descartados_possui_site']}")
        print(f"Total descartado por avaliacoes acima do limite: {stats['descartados_avaliacoes']}")
        print(f"Total aceito: {stats['leads_raspagem']}")
        print(f"Total novo adicionado: {stats['novos_adicionados']}")
        print(f"Total duplicado: {stats['duplicados_ignorados']}")
        print(f"Total geral da base: {stats['total_geral_base']}")

        if stats["mensagem_exportacao"]:
            print(stats["mensagem_exportacao"])

    if retornar_estatisticas:
        return enriched_leads, stats

    return enriched_leads


def main() -> None:
    print("LeadMaps - Pesquisa empresas no Google Maps e exporta para Excel")
    print("-" * 62)

    niche = input("Digite o nicho: ").strip()
    city = input("Digite a cidade: ").strip()

    if not niche or not city:
        print("Nicho e cidade sao obrigatorios.")
        return

    limit = input("Quantidade de resultados para analisar [100]: ").strip()
    limit = int(limit) if limit.isdigit() else MAX_COMPANIES
    max_reviews = input("Avaliacoes maximas [sem limite]: ").strip()
    max_reviews = int(max_reviews) if max_reviews.isdigit() else None

    print("\nBuscando leads...")
    companies = buscar_leads(niche, city, limit, True, True, max_reviews)
    print(f"\n{len(companies)} leads exportados para {EXPORT_PATH}")


if __name__ == "__main__":
    main()
